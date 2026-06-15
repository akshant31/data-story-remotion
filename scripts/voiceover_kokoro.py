#!/usr/bin/env python3
"""
FREE / LOCAL voiceover with Kokoro (runs on CPU, no API key, commercial-OK).

Setup (one time):
    pip install kokoro soundfile openpyxl numpy
    # Kokoro also needs espeak-ng:  apt install espeak-ng   (or: brew install espeak-ng)

Run:
    python scripts/voiceover_kokoro.py reels-2weeks.xlsx
    # optional voice:  VO_VOICE=af_heart  (others: am_adam, bf_emma, af_bella, ...)

Writes public/vo/<reel>/NN.wav + timing.json, exactly like the cloud script,
so `npm run reels` picks the audio up automatically.
"""
import sys, os, re, json
import numpy as np
import soundfile as sf
from openpyxl import load_workbook
from kokoro import KPipeline

xlsx = sys.argv[1] if len(sys.argv) > 1 else "reels-2weeks.xlsx"
VOICE = os.environ.get("VO_VOICE", "af_heart")
pipeline = KPipeline(lang_code="a")  # 'a' = American English; 'b' = British

wb = load_workbook(xlsx)
ws = wb["Reels"]
hdr = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
col = {h: i + 1 for i, h in enumerate(hdr)}


def cell(r, name):
    v = ws.cell(r, col[name]).value
    return "" if v is None else str(v).strip()


reels = {}
for r in range(3, ws.max_row + 1):
    name, typ = cell(r, "Reel"), cell(r, "Type")
    if not name or not typ:
        continue
    reels.setdefault(name, []).append(cell(r, "Narration"))

slug = lambda s: re.sub(r"[^a-z0-9-_]+", "-", s.lower())

for name, lines in reels.items():
    d = f"public/vo/{slug(name)}"
    os.makedirs(d, exist_ok=True)
    timing = []
    for i, text in enumerate(lines, start=1):
        if not text:
            timing.append({"index": i, "file": None, "text": "", "seconds": 0})
            continue
        audio = np.concatenate([a for _, _, a in pipeline(text, voice=VOICE)])
        path = f"{d}/{i:02d}.wav"
        sf.write(path, audio, 24000)
        secs = round(len(audio) / 24000, 2)
        timing.append({"index": i, "file": f"{i:02d}.wav", "text": text, "seconds": secs})
        print(f"  {name} {i:02d}  {secs}s  -> {path}")
    json.dump({"clips": timing}, open(f"{d}/timing.json", "w"), indent=2)

print("\nDone. Now run:  npm run reels")
